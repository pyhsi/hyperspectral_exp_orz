# coding: utf-8
import scipy.io as sio
import numpy as np

def get_data(data_file):
    """
    Load training dataset and test dataset in the data_file

    Arg:
        data_file: File of the original data

    Returns:
        train_data: Training dataset
        train_label: The label of training data
        train_pred: The prediction of training data
        test_data: Test dataset
        test_label: the label of test data
        test_pred: The prediction of test data

    """

    data = sio.loadmat(data_file)
    train_data = data['train_data']
    train_label = data['train_label']
    train_pred = data['train_prediction']
    test_data = data['test_data']
    test_label = data['test_label']
    test_pred = data['test_prediction']

    return train_data, train_label, train_pred, test_data, test_label, test_pred

from openpyxl import Workbook
from openpyxl import load_workbook

def get_confuse_matrix(label_gt, label_pred):
    """
    Calculating the confuse matrix.

    Args:
        label_gt:
        label_pred:

    Return:
        confuse_matrix: One element [i, j] means the sample numbers of  i-th class predicting to j-th class, the diagonal elements means the correction of the predictiong numbers.

    """
    label_min = np.min(label_gt)
    label_max = np.max(label_gt)
    label_num = label_max - label_min + 1
    confuse_matrix = np.zeros([label_num, label_num], int)

    for idx in range(np.size(label_gt, 1)):
        l_gt = label_gt[0, idx]
        l_pred = label_pred[0, idx]
        confuse_matrix[l_gt, l_pred] += 1

    print('Confuse matrix done.')
    return confuse_matrix

def confuse_save(conf_matrix, sheet_name, file_name, new_flag):
    """
    Save confuse matrix into excel file.

    Args:
        conf_matrix: Confuse matrix from get_confuse_matrix()
        sheet_name: Define sheet name
        file_name: file dir
        new_flag: Whether write to new workbook

    """
    label_num = np.size(conf_matrix, 1)

    if new_flag:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    else:
        wb = load_workbook(file_name)
        ws = wb.create_sheet(title = sheet_name)

    for idx1 in range(1, label_num + 1):
        for idx2 in range(1, label_num + 1):
             ws.cell(row = idx1, column = idx2, value = conf_matrix[idx1 - 1, idx2 - 1])

    wb.save(filename = file_name)
    print('Save the matrix.')

def get_misClassify_neighbors_info(conf_matrix):
    """
    Get the misclassified samples' information according to the confuse matrix.
    Information including a sample's position, class and its 8-neighbors' class.

    Arg:
        conf_matrix: Confuse matrix from get_confuse_matrix()

    Returns:
        mis_class: Tuple, [Misclassified sample's class, prediction].
        mis_position: Misclassified sample's position.
        mis_neighbors: The class of Misclassified sample's 8-neighbors.

    """
    m, n = np.size(conf_matrix)


if __name__ == '__main__':
    root = 'F:\hsi_result\original'
    for i in [1, 4, 8]:
        train_data, train_label, train_pred, test_data, test_label, test_pred = get_data(root + '\KSC\data\data' + str(i) + '.mat')
        confuse_matrix_train = get_confuse_matrix(train_label, train_pred)
        confuse_matrix_test = get_confuse_matrix(test_label, test_pred)
        confuse_save(confuse_matrix_train, 'train_' + str(i), root, 1)
        confuse_save(confuse_matrix_test, 'test_' + str(i), root, 1)

